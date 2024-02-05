import csv
import os
from pypdf import PdfReader
from openpyxl import load_workbook


def test_users_csv(create_archive):
    path = os.path.join(create_archive, 'file3.csv')
    with open(path) as csvfile:
        reader = csv.DictReader(csvfile, delimiter=';')
        names = [row['mama'] for row in reader]
    assert 'papa' in names


def test_xlsx_file(create_archive):
    path = os.path.join(create_archive, 'file2.xlsx')
    open_xlsx = load_workbook(path)
    sheet = open_xlsx.active
    name = sheet.cell(row=3, column=7).value
    assert name == 'August'


def test_pdf_page_content(create_archive):
    path = os.path.join(create_archive, 'file1.pdf')
    text_to_search = "Тестирование демонстрирует наличие дефектов."
    with open(path, 'rb') as file:
        reader = PdfReader(file)
        text_found = any(text_to_search in page.extract_text() for page in reader.pages)
        assert text_found, "Текст не найден в PDF файле."